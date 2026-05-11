from fastapi import APIRouter,Query, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse, JSONResponse

from sqlalchemy.orm import Session
from sqlalchemy import select

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment

from collections import defaultdict 
from typing import List,Optional
from datetime import datetime
import pandas as pd
import io

from ..auth.security   import get_current_user
from ..promotion.full_import_service import FullImportService
from ..database.common.connet_database_postgres import get_db
from ..database.models.views import vwsummaryfileimport,vwfileimport,VwExportFile
from ..database.models.postgres_models import MaUser, MInfoImportFile, MInfoImportFile
from ..version.all_schemas  import InfoImportCreate,BaseModel
from ..common.minio_service import MinioService


router = APIRouter(prefix="/upload", tags=["import file Control"])
try:
    minio_service = MinioService()
except:
    minio_service = None
class FileInfoResponse(BaseModel):
    id: int
    file_name: Optional[str]
    sheet: Optional[str]
    status: Optional[int]
    r_row: Optional[int]
    w_row: Optional[int]
    description: Optional[str]
    user_mk: Optional[str]
    name: Optional[str]

    class Config:
        from_attributes = True

class promotion_id(BaseModel):
    entity_code: List[str] = ['']   
# ==========================================
# 2. Schema สำหรับ VW_SUMMARY_FILE_IMPORT
# ==========================================
class FileSummaryResponse(BaseModel):
    version_id: int=0
    title: Optional[str]=None
    description: Optional[str]=None
    file_name: Optional[int]=None
    sheet: Optional[int]=None
    r_row: Optional[int]=None
    ww_row: Optional[int]=None
    read_row: Optional[int]=None
    # name: Optional[str]=None
    data: Optional[List[FileInfoResponse]] = None    
    class Config:
        from_attributes = True

TABEL_LINE_LEFT = "thin"
TABEL_LINE_RIGHT = "thin"
TABEL_LINE_TOP = "thin"
TABEL_LINE_BOTTOM = "thin"
FONT_BARCODE = "Bar-Code 39"          # ชื่อฟอนต์ Barcode (สมมติ)
FONT_NOMAL = "Anuphan"
BACKGROUND_FONT_NOMAL = "Anuphan"
HEARDER_FONT_SIZE = 30
BACKGROUND_FONT_SIZE = 12
BACKGROUND_COLOR = "000000"      # สีตัวอักษร Header
HEARDER_COLOR = "DDDDDD"         # สีพื้นหลัง Header
SOLID = "solid"
BACKGROUND_WRAP_TEXT = True
BACKGROUND_HORIZONTAL = "center"
BACKGROUND_VERTICAL = "center"
COLUMN_BARCODE_NAMES = ['BARCODE CODE39', 'COUPON CODE39'] 
PRINT_SIZE_LEFT = 0.25     #1.9
PRINT_SIZE_RIGHT = 0.25     #1.9
PRINT_SIZE_TOP = 0.75     #2.5
PRINT_SIZE_BOTTOM = 0.75     #2.5
PRINT_SIZE_HEADER = 0.30     #1.3
PRINT_SIZE_FOOTER = 0.30     #1.3
PRINT_WIDTH = 1
PRINT_HEIGT = 0
PRINT_SCALE = 100
WIDTH_COLUMN=[10,10,6.3,6.3,7.5,5.3,5.3,5.3,11,11,16.5,53,7.5,10,50,15.3,53,8.6,15,15,15,15]
PRINT_TEXT_HEARDER_RIGHT = "&[Page] of &[Pages]"
PRINT_TEXT_HEARDER_CENTER = "&[Tab]"


# class FileUploadReportResponse(BaseModel):
#     version_id_filtered: int
#     total_files: int
#     total_summaries: int
#     file_information: List[FileInfoResponse]
#     summary_list: List[FileSummaryResponse]

# @router.post("/insert")
# async def insert_info_import(data: InfoImportCreate, db: Session = Depends(get_db)):
#     try:
#         new_info = MInfoImportFile(
#             v_id=data.v_id,
#             status=data.status,
#             description=data.description,
#             user_create=data.user_create,
#             date_create=datetime.utcnow()
#         )
#         db.add(new_info)
#         db.commit()
#         db.refresh(new_info)
        
#         # Return id กลับไปหา User ตามคำสั่ง
#         return {"id": new_info.id, "message": "Insert successful"}
#     except Exception as e:
#         db.rollback()
#         print(str(e))
#         raise HTTPException(status_code=400, detail=str(e))

# @router.post("/upload_data", response_model=ImportResponse)
# def execute_import(info: ImportInformation, db: Session = Depends(get_db)):
#     """
#     API สำหรับเริ่มกระบวนการ Import โดยรับข้อมูลเป็น JSON Object (ImportInformation)
#     """
#     try:
#         service = PromotionImportService(db)
#         service.process_import(info) # type: ignore
        
#         return {
#             "status": "Success", 
#             "message": "Import completed successfully",
#             "version_id": info.version_id,
#             "processed_sheets": info.sheet
#         }
#     except Exception as e:
#         raise HTTPException(status_code=500, detail=str(e))
   

@router.post("/fileexcel")
async def upload_and_import(
    version_id: int = Form(..., description="เลข Version id "),
    current_user: MaUser = Depends(get_current_user),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    API เดียวจบ: รับไฟล์ -> ลง MinIO -> อ่าน -> ลง Database Postgres
    """
    try:
        file_content = await file.read()
        
        service = FullImportService(db)

        result = service.run_process(
            file_content=file_content,
            filename=file.filename, # type: ignore
            version=version_id,
            user_id=current_user.user_id)
        
        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")
    
@router.post("/insert")
async def insert_info_import(data: InfoImportCreate, 
                            current_user: MaUser = Depends(get_current_user),
                            db: Session = Depends(get_db)):
    try:
        new_info = MInfoImportFile(
            v_id=data.version_id,
            status=data.status,
            description=data.description,
            user_create=current_user.user_id,
            date_create=datetime.utcnow()
        )
        db.add(new_info)
        db.commit()
        db.refresh(new_info)
        return {"id": new_info.id, "message": "Insert successful"}
    except Exception as e:
        db.rollback()
        print(str(e))
        raise HTTPException(status_code=400, detail=str(e))



@router.get("/fileinformation", response_model=List[FileSummaryResponse])
async def get_file_upload_report(
    version_id: int = 0,
    skip: int = 0, 
    limit: int = 100, 
    db: Session = Depends(get_db)
):
    """
    ดึงรายงานข้อมูลสรุปการนำเข้าไฟล์ พร้อมรายละเอียดไฟล์ที่ซ้อนอยู่ด้านใน (Nested)
    """
    stmt_info = select(vwfileimport)
    if version_id != 0:
        stmt_info = stmt_info.where(vwfileimport.version_no == version_id)
    file_infos = db.scalars(stmt_info).all()  
    info_dict = defaultdict(list)
    for info in file_infos:
        info_dict[info.version_id].append(info)
    stmt_summary = select(vwsummaryfileimport)
    if version_id != 0:
        stmt_summary = stmt_summary.where(vwsummaryfileimport.version_id == version_id)
    stmt_summary = stmt_summary.offset(skip).limit(limit)
    summaries = db.scalars(stmt_summary).all()  
    response: List[FileSummaryResponse] = []
    print(info_dict.items())
    for summary in summaries:
        print(summary.version_no)
        summary_dict = {
            "version_id": summary.version_no,
            "title": getattr(summary, 'title', None),
            "description": getattr(summary, 'description', None),
            "file_name": getattr(summary, 'file_name', None),
            "sheet": getattr(summary, 'sheet', None),
            "r_row": getattr(summary, 'r_row', None),
            "ww_row": getattr(summary, 'ww_row', None),
            "read_row": getattr(summary, 'read_row', None),
            "data": info_dict.get(summary.version_no, []) 
        }
        response.append(FileSummaryResponse(**summary_dict))

    return response

@router.put("/update-status/{data}")
async def update_info_status(data, db: Session = Depends(get_db)):
    info_record = db.query(MInfoImportFile).filter(MInfoImportFile.id == data).first()
    if not info_record:
        raise HTTPException(status_code=404, detail="Record not found")
    try:
        info_record.status = 3
        db.commit()
        return {"message": "Update status successful"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))





@router.get("/export")
def get_export_files(
    versionid: Optional[str] = Query(None, description="รหัส Version ID"),
    fileid: Optional[str] = Query(None, description="รหัส File ID"),
    export_pdf: Optional[bool] = Query(False, description="....."),
    db: Session = Depends(get_db)
):
    """
    API สำหรับดึงข้อมูลจาก vw_export_file
    รองรับการค้นหาด้วย versionid และ fileid (หากไม่ส่งมา จะดึงข้อมูลทั้งหมด)
    """
    def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    # เติมค่าว่างด้วย string เปล่า และลบช่องว่างส่วนเกิน
        df = df.fillna("")
        df = df.map(lambda x: str(x).strip() if isinstance(x, str) else x)
        return df

    try:
        # 1. สร้าง Base Query ตั้งต้น
        query = db.query(VwExportFile)
        if versionid:
            query = query.filter(VwExportFile.version_id == versionid)
            
        if fileid:
            query = query.filter(VwExportFile.file_id == fileid)

  
        df = pd.read_sql(query.statement, query.session.get_bind())
        print(df.shape)
        df = clean_data(df)
        if df.empty:
            return JSONResponse(status_code=403, content={"success": False, "message": "No data found."})
        date_columns = ['Active From', 'Active To', 'OPTIMAL_DATE']

        for col in date_columns:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')
                
                df[col] = df[col].apply(
                    lambda x: x.strftime('%d/%m/%y') if pd.notnull(x) else ''
                )

        df = df.astype(str)
        
        sort_cols = [c for c in ["OPTIMAL_DATE", "WORKSHEET"] if c in df.columns]
        if sort_cols:
            df = df.sort_values(by=sort_cols)

        output = io.BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Export_Data" # type: ignore

        # 3. Define Styles
        thin_border = Border(
            left=Side(style=TABEL_LINE_LEFT),
            right=Side(style=TABEL_LINE_RIGHT),
            top=Side(style=TABEL_LINE_TOP),
            bottom=Side(style=TABEL_LINE_BOTTOM)
        )
        thick_top_border = Border(
            top=Side(style='thick'),
            left=Side(style=TABEL_LINE_LEFT),
            right=Side(style=TABEL_LINE_RIGHT),
            bottom=Side(style=TABEL_LINE_BOTTOM)
        )
        
        barcode_font = Font(name=FONT_BARCODE, size=HEARDER_FONT_SIZE)
        default_font = Font(name=FONT_NOMAL, size=BACKGROUND_FONT_SIZE)
        header_font = Font(
            name=BACKGROUND_FONT_NOMAL, 
            color=BACKGROUND_COLOR, 
            bold=True, 
            size=BACKGROUND_FONT_SIZE
        )
        
        light_grey_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        header_fill = PatternFill(start_color=HEARDER_COLOR, end_color=HEARDER_COLOR, fill_type=SOLID)
        alignment_style = Alignment(wrap_text=BACKGROUND_WRAP_TEXT, horizontal=BACKGROUND_HORIZONTAL, vertical=BACKGROUND_VERTICAL)

        # 4. Write Data (เขียนครั้งเดียวต่อยาวลงมา)
        headers = list(df.columns)
        
        # เขียน Header ที่บรรทัดที่ 1
        worksheet.append(headers) # type: ignore

        # เขียนข้อมูลทั้งหมด (ต่อยาวลงมา)
        for row in df.itertuples(index=False):
            worksheet.append(list(row)) # type: ignore

        # 5. Formatting Logic (ทำทีเดียวทั้ง Sheet)
        col_map = {name: i+1 for i, name in enumerate(headers)}
        att_mode_idx = col_map.get("AttachmentMode")
        barcode_col_indices = [col_map[c] for c in COLUMN_BARCODE_NAMES if c in col_map]
        
        prev_first_col_val = None
        max_col = worksheet.max_column # type: ignore

        # เริ่ม Loop ตั้งแต่แถว 2 จนจบ
        for r_idx, row_cells in enumerate(worksheet.iter_rows(min_row=2, max_col=max_col), start=2): # type: ignore
            curr_first_col_val = row_cells[0].value
            
            # เช็ค Attachment Mode (สีพื้นหลัง)
            is_exclude = False
            if att_mode_idx:
                att_val = row_cells[att_mode_idx - 1].value
                is_exclude = str(att_val).strip().upper() not in ["INCLUDE",""]

            # Logic เส้นขอบ
            current_border = thin_border
            if r_idx == 2 or curr_first_col_val != prev_first_col_val:
                current_border = thick_top_border
            
            prev_first_col_val = curr_first_col_val

            # Apply Style ราย Cell
            for c_idx, cell in enumerate(row_cells):
                cell_col_num = c_idx + 1
                
                cell.font = default_font
                cell.border = current_border
                cell.alignment = alignment_style

                if is_exclude:
                    cell.fill = light_grey_fill

                if cell_col_num in barcode_col_indices:
                    cell.font = barcode_font

        # 6. Header Style & Widths
        for i, col_name in enumerate(headers):
            cell = worksheet.cell(row=1, column=i+1)  # type: ignore
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = alignment_style
            
            if i < len(WIDTH_COLUMN):
                worksheet.column_dimensions[get_column_letter(i+1)].width = WIDTH_COLUMN[i]  # type: ignore
            else:
                worksheet.column_dimensions[get_column_letter(i+1)].width = 15  # type: ignore

        # 7. Page Setup
        worksheet.row_dimensions[1].height = 105  # type: ignore
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3  # type: ignore
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE  # type: ignore
        worksheet.print_title_rows = "1:1"  # type: ignore
        worksheet.oddHeader.center.text = PRINT_TEXT_HEARDER_CENTER  # type: ignore
        worksheet.oddHeader.right.text = PRINT_TEXT_HEARDER_RIGHT  # type: ignore

        # 8. Save & Return
        workbook.save(output)
        output.seek(0)
        filename = f"Promotion_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        headers_http = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
        
        return StreamingResponse(
            io.BytesIO(output.getvalue()), 
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
            headers=headers_http
        )

    except Exception as e:
        print(f"Export Error: {e}")
        return JSONResponse(status_code=500, content={"success": False, "message": str(e)})



# import io
# import pandas as pd
# import openpyxl
# from datetime import datetime
# from typing import Optional
# from fastapi import APIRouter, Depends, Query
# from fastapi.responses import JSONResponse, StreamingResponse
# from sqlalchemy.orm import Session

# # --- Excel Imports ---
# from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
# from openpyxl.utils import get_column_letter

# # --- PDF Imports ---
# from reportlab.lib.pagesizes import A3, landscape
# from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
# from reportlab.lib import colors
# from reportlab.pdfgen import canvas
# from pypdf import PdfReader, PdfWriter
# from reportlab.pdfbase import pdfmetrics
# from reportlab.pdfbase.ttfonts import TTFont

# import os
# # --- Constants ของคุณ (คงไว้เหมือนเดิม) ---
# TABEL_LINE_LEFT = "thin"
# TABEL_LINE_RIGHT = "thin"
# TABEL_LINE_TOP = "thin"
# TABEL_LINE_BOTTOM = "thin"
# FONT_BARCODE = "Bar-Code 39"          
# FONT_NOMAL = "Anuphan"
# BACKGROUND_FONT_NOMAL = "Anuphan"
# HEARDER_FONT_SIZE = 30
# BACKGROUND_FONT_SIZE = 12
# BACKGROUND_COLOR = "000000"      
# HEARDER_COLOR = "DDDDDD"         
# SOLID = "solid"
# BACKGROUND_WRAP_TEXT = True
# BACKGROUND_HORIZONTAL = "center"
# BACKGROUND_VERTICAL = "center"
# COLUMN_BARCODE_NAMES = ['BARCODE CODE39', 'COUPON CODE39'] 
# PRINT_SIZE_LEFT = 0.25    
# PRINT_SIZE_RIGHT = 0.25    
# PRINT_SIZE_TOP = 0.75    
# PRINT_SIZE_BOTTOM = 0.75    
# PRINT_SIZE_HEADER = 0.30    
# PRINT_SIZE_FOOTER = 0.30    
# PRINT_WIDTH = 1
# PRINT_HEIGT = 0
# PRINT_SCALE = 100
# WIDTH_COLUMN = [10,10,6.3,6.3,7.5,5.3,5.3,5.3,11,11,16.5,53,7.5,10,50,15.3,53,8.6,15,15,15,15]
# PRINT_TEXT_HEARDER_RIGHT = "&[Page] of &[Pages]"
# PRINT_TEXT_HEARDER_CENTER = "&[Tab]"
# try:
#     pdfmetrics.registerFont(TTFont('THSarabun', 'THSarabunNew.ttf'))
#     pdfmetrics.registerFont(TTFont('THSarabun-Bold', 'THSarabunNew Bold.ttf'))
#     PDF_FONT_NORMAL = 'THSarabun'
#     PDF_FONT_BOLD = 'THSarabun-Bold'
#     PDF_FONT_SIZE_DATA = 12
#     PDF_FONT_SIZE_HEADER = 14
# except Exception as e:
#     print(f"Warning: Thai font not found. Using default. Error: {e}")
#     PDF_FONT_NORMAL = 'Helvetica'
#     PDF_FONT_BOLD = 'Helvetica-Bold'
#     PDF_FONT_SIZE_DATA = 8
#     PDF_FONT_SIZE_HEADER = 9

# # ถ้ามี Barcode Font ให้ลงทะเบียนด้วย (ถ้าไม่มีให้ลบออกได้)
# try:
#     pdfmetrics.registerFont(TTFont(FONT_BARCODE, 'barcode39.ttf'))
# except:
#     pass
# # ==========================================
# # 1. Custom Canvas สำหรับ PDF (Group Header/Footer)
# # ==========================================
# class GroupCanvas(canvas.Canvas):
#     def __init__(self, *args, **kwargs):
#         canvas.Canvas.__init__(self, *args, **kwargs)
#         self._saved_page_states = []
#         self.group_info = ("Unknown", "Unknown", "Unknown")

#     def showPage(self):
#         self._saved_page_states.append(dict(self.__dict__))
#         self._startPage() # type: ignore

#     def save(self):
#         num_pages = len(self._saved_page_states)
#         for state in self._saved_page_states:
#             self.__dict__.update(state)
#             self.draw_header_footer(num_pages)
#             canvas.Canvas.showPage(self)
#         canvas.Canvas.save(self)

#     def draw_header_footer(self, page_count):
#         # ใช้ฟอนต์ไทยที่โหลดมา
#         self.setFont(PDF_FONT_BOLD, 16) 
#         sheet, worksheet, active_from = self.group_info
        
#         header_text = f"SHEET: {sheet}   |   WORKSHEET: {worksheet}   |   Active From: {active_from}"
#         page_width, page_height = landscape(A3)
#         self.drawCentredString(page_width / 2, page_height - 30, header_text)

#         self.setFont(PDF_FONT_NORMAL, 12)
#         self.drawCentredString(page_width / 2, 20, f"Page {self._pageNumber} of {page_count}") # type: ignore

# # ==========================================
# # 2. API Function หลัก
# # ==========================================
# @router.get("/export-files") # ปรับ Path ตามของคุณ
# def get_export(
#     versionid: Optional[str] = Query(None, description="รหัส Version ID"),
#     fileid: Optional[str] = Query(None, description="รหัส File ID"),
#     export_pdf: Optional[bool] = Query(False, description="True=PDF, False=Excel"),
#     db: Session = Depends(get_db)
# ):
#     def clean_data(df: pd.DataFrame) -> pd.DataFrame:
#         str_cols = df.select_dtypes(include=['object', 'string']).columns
#         for col in str_cols:
#             try: df[col] = df[col].str.strip()
#             except AttributeError: pass
#         return df.fillna("")

#     try:
#         query = db.query(VwExportFile)
#         if versionid: query = query.filter(VwExportFile.version_id == versionid)
#         if fileid: query = query.filter(VwExportFile.file_id == fileid)

#         df = pd.read_sql(query.statement, query.session.get_bind())
    
#         if df.empty:
#             return JSONResponse(status_code=403, content={"success": False, "message": "No data found."})
            
#         df = clean_data(df)
        
#         date_columns = ['Active From', 'Active To', 'OPTIMAL_DATE']
#         for col in date_columns:
#             if col in df.columns:
#                 df[col] = pd.to_datetime(df[col], errors='coerce')
#                 df[col] = df[col].dt.strftime('%d/%m/%y').fillna('') # type: ignore

#         df = df.astype(str)
#         sort_cols = [c for c in ["OPTIMAL_DATE", "WORKSHEET"] if c in df.columns]
#         if sort_cols: df = df.sort_values(by=sort_cols)

#         headers = list(df.columns)
#         col_map = {name: i for i, name in enumerate(headers)} # 0-indexed for list
#         att_mode_idx = col_map.get("AttachmentMode")
#         barcode_col_indices = {col_map[c] for c in COLUMN_BARCODE_NAMES if c in col_map}

#         # ==========================================
#         # 🟢 ออกเอกสาร PDF
#         # ==========================================
#         if export_pdf:
#             pdf_writer = PdfWriter()
#             page_width, page_height = landscape(A3)
            
#             # กำหนดขอบกระดาษ (Margin)
#             pdf_margin = 40
#             printable_width = page_width - (pdf_margin * 2)

#             groups = df.groupby(['SHEET', 'WORKSHEET', 'Active From'], dropna=False)

#             # [✨ ฟีเจอร์ใหม่] คำนวณความกว้างคอลัมน์ให้พอดีหน้ากระดาษเป๊ะๆ (Proportional Width)
#             total_excel_width = 0
#             base_widths = []
#             for i in range(len(headers)):
#                 w = WIDTH_COLUMN[i] if i < len(WIDTH_COLUMN) else 15
#                 base_widths.append(w)
#                 total_excel_width += w
                
#             pdf_col_widths = [(w / total_excel_width) * printable_width for w in base_widths]

#             for group_name, group_df in groups:
#                 sheet, worksheet, active_from = group_name
#                 group_buffer = io.BytesIO()
                
#                 # กำหนด Margin ของเอกสารให้ตรงกับที่คำนวณไว้
#                 doc = SimpleDocTemplate(
#                     group_buffer, 
#                     pagesize=landscape(A3), 
#                     leftMargin=pdf_margin, 
#                     rightMargin=pdf_margin,
#                     topMargin=50, 
#                     bottomMargin=50
#                 )

#                 table_data = [headers] + group_df.values.tolist()
                
#                 # 1. จัด Row Heights
#                 row_heights = [40] # ความสูง Header
#                 for row in group_df.values.tolist():
#                     att_val = row[att_mode_idx] if att_mode_idx is not None else ""
#                     row_heights.append(30 if str(att_val).strip().upper() == "INCLUDE" else 20)

#                 # 2. Base Table Styles (ใช้ฟอนต์ไทย)
#                 pdf_valign = 'MIDDLE' if BACKGROUND_VERTICAL.strip().upper() == 'CENTER' else BACKGROUND_VERTICAL.upper()

#                 pdf_styles = [
#                     ('ALIGN', (0,0), (-1,-1), BACKGROUND_HORIZONTAL.upper()),
#                     ('VALIGN', (0,0), (-1,-1), pdf_valign),
#                     ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#' + HEARDER_COLOR)),
#                     ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#' + BACKGROUND_COLOR)),
                    
#                     # ตั้งค่า Font มาตรฐานทั้งตาราง
#                     ('FONTNAME', (0,0), (-1,0), PDF_FONT_BOLD),
#                     ('FONTSIZE', (0,0), (-1,0), PDF_FONT_SIZE_HEADER),
#                     ('FONTNAME', (0,1), (-1,-1), PDF_FONT_NORMAL),
#                     ('FONTSIZE', (0,1), (-1,-1), PDF_FONT_SIZE_DATA),
                    
#                     ('GRID', (0,0), (-1,-1), 0.5, colors.black),
#                 ]

#                 # 3. Dynamic Styles 
#                 prev_first_col_val = None
#                 for r_idx, row in enumerate(group_df.values.tolist(), start=1):
#                     curr_first_col_val = row[0]
                    
#                     # สีพื้นหลัง Exclude
#                     att_val = row[att_mode_idx] if att_mode_idx is not None else ""
#                     if str(att_val).strip().upper() not in ["INCLUDE", ""]:
#                         pdf_styles.append(('BACKGROUND', (0, r_idx), (-1, r_idx), colors.HexColor('#E0E0E0')))

#                     # เส้นขอบหนา
#                     if r_idx == 1 or curr_first_col_val != prev_first_col_val:
#                         pdf_styles.append(('LINEABOVE', (0, r_idx), (-1, r_idx), 1.5, colors.black))
#                     prev_first_col_val = curr_first_col_val

#                     # เปลี่ยน Font เป็น Barcode สำหรับคอลัมน์ Barcode
#                     for c_idx in barcode_col_indices:
#                         pdf_styles.append(('FONTNAME', (c_idx, r_idx), (c_idx, r_idx), FONT_BARCODE))
#                         pdf_styles.append(('FONTSIZE', (c_idx, r_idx), (c_idx, r_idx), HEARDER_FONT_SIZE))

#                 t = Table(table_data, colWidths=pdf_col_widths, rowHeights=row_heights, repeatRows=1)
#                 t.setStyle(TableStyle(pdf_styles))

#                 def make_canvas(*args, **kwargs):
#                     c = GroupCanvas(*args, **kwargs)
#                     c.group_info = (sheet, worksheet, active_from) # type: ignore
#                     return c

#                 doc.build([t], canvasmaker=make_canvas)
#                 group_buffer.seek(0)

#                 # นำมาต่อกันและเช็คหน้าว่าง
#                 group_reader = PdfReader(group_buffer)
#                 num_pages = len(group_reader.pages)
#                 for page in group_reader.pages:
#                     pdf_writer.add_page(page)

#                 if num_pages % 2 != 0:
#                     pdf_writer.add_blank_page(width=page_width, height=page_height)

#             final_pdf_buffer = io.BytesIO()
#             pdf_writer.write(final_pdf_buffer)
#             final_pdf_buffer.seek(0)
#             filename = f"Promotion_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
#             return StreamingResponse(
#                 final_pdf_buffer, 
#                 media_type='application/pdf', 
#                 headers={'Content-Disposition': f'attachment; filename="{filename}"'}
#             )
#         # ==========================================
#         # 🔵 ออกเอกสาร Excel (โค้ดเดิมของคุณ)
#         # ==========================================
#         else:
#             output = io.BytesIO()
#             workbook = openpyxl.Workbook()
#             worksheet = workbook.active
#             worksheet.title = "Export_Data" # type: ignore

#             thin_border = Border(left=Side(style=TABEL_LINE_LEFT), right=Side(style=TABEL_LINE_RIGHT), top=Side(style=TABEL_LINE_TOP), bottom=Side(style=TABEL_LINE_BOTTOM))
#             thick_top_border = Border(top=Side(style='thick'), left=Side(style=TABEL_LINE_LEFT), right=Side(style=TABEL_LINE_RIGHT), bottom=Side(style=TABEL_LINE_BOTTOM))
            
#             barcode_font = Font(name=FONT_BARCODE, size=HEARDER_FONT_SIZE)
#             default_font = Font(name=FONT_NOMAL, size=BACKGROUND_FONT_SIZE)
#             header_font = Font(name=BACKGROUND_FONT_NOMAL, color=BACKGROUND_COLOR, bold=True, size=BACKGROUND_FONT_SIZE)
            
#             light_grey_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
#             header_fill = PatternFill(start_color=HEARDER_COLOR, end_color=HEARDER_COLOR, fill_type=SOLID)
#             alignment_style = Alignment(wrap_text=BACKGROUND_WRAP_TEXT, horizontal=BACKGROUND_HORIZONTAL, vertical=BACKGROUND_VERTICAL)

#             worksheet.append(headers) # type: ignore
#             for row in df.itertuples(index=False):
#                 worksheet.append(list(row)) # type: ignore

#             col_map_excel = {name: i+1 for i, name in enumerate(headers)} # 1-indexed for openpyxl
#             att_mode_idx_ex = col_map_excel.get("AttachmentMode")
#             barcode_col_indices_ex = [col_map_excel[c] for c in COLUMN_BARCODE_NAMES if c in col_map_excel]
            
#             prev_first_col_val = None
#             max_col = worksheet.max_column # type: ignore

#             for r_idx, row_cells in enumerate(worksheet.iter_rows(min_row=2, max_col=max_col), start=2):  # type: ignore
#                 curr_first_col_val = row_cells[0].value
                
#                 is_exclude = False
#                 is_include = False
#                 if att_mode_idx_ex:
#                     att_val = row_cells[att_mode_idx_ex - 1].value
#                     att_val_str = str(att_val).strip().upper()
#                     if att_val_str == "INCLUDE": is_include = True
#                     elif att_val_str != "": is_exclude = True

#                 # เพิ่มฟีเจอร์ปรับ Height สำหรับ INCLUDE
#                 if not is_include:
#                     worksheet.row_dimensions[r_idx].height = 30 # type: ignore

#                 current_border = thin_border
#                 if r_idx == 2 or curr_first_col_val != prev_first_col_val:
#                     current_border = thick_top_border
#                 prev_first_col_val = curr_first_col_val

#                 for c_idx, cell in enumerate(row_cells):
#                     cell_col_num = c_idx + 1
#                     cell.font = default_font
#                     cell.border = current_border
#                     cell.alignment = alignment_style

#                     if is_exclude: cell.fill = light_grey_fill
#                     if cell_col_num in barcode_col_indices_ex: cell.font = barcode_font

#             for i, col_name in enumerate(headers):
#                 cell = worksheet.cell(row=1, column=i+1)  # type: ignore
#                 cell.font = header_font
#                 cell.fill = header_fill
#                 cell.border = thin_border
#                 cell.alignment = alignment_style
#                 if i < len(WIDTH_COLUMN):
#                     worksheet.column_dimensions[get_column_letter(i+1)].width = WIDTH_COLUMN[i]  # type: ignore
#                 else:
#                     worksheet.column_dimensions[get_column_letter(i+1)].width = 15 # type: ignore

#             worksheet.row_dimensions[1].height = 105  # type: ignore
#             worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3  # type: ignore
#             worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE  # type: ignore
#             worksheet.print_title_rows = "1:1"  # type: ignore
#             worksheet.oddHeader.center.text = PRINT_TEXT_HEARDER_CENTER  # type: ignore
#             worksheet.oddHeader.right.text = PRINT_TEXT_HEARDER_RIGHT  # type: ignore

#             workbook.save(output)
#             output.seek(0)
#             filename = f"Promotion_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
#             return StreamingResponse(
#                 io.BytesIO(output.getvalue()), 
#                 media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
#                 headers={'Content-Disposition': f'attachment; filename="{filename}"'}
#             )

#     except Exception as e:
#         print(f"Export Error: {e}")
#         return JSONResponse(status_code=500, content={"success": False, "message": str(e)})