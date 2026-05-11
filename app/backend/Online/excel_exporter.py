import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Dict, Any

class ExcelExporter:
    @staticmethod
    def generate_excel_bytes(fetch_result: Dict[str, Any]) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Config"
        
        header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        column_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
        
        row_idx = 1
        
        # ดึงข้อมูลจากคีย์ "data" ซึ่งเป็น List
        tables_list = fetch_result.get("data", [])
        
        for table_dict in tables_list:
            for table_name, rows_data in table_dict.items():
                ws.cell(row=row_idx, column=1, value=table_name)
                ws.cell(row=row_idx, column=1).fill = header_fill
                ws.cell(row=row_idx, column=1).font = Font(bold=True)
                row_idx += 2
                
                # ตรวจสอบว่ามีข้อมูลใน Table นี้หรือไม่
                if rows_data:
                    # pandas รองรับการนำ List of Dictionaries มาสร้าง DataFrame ได้เลย (คอลัมน์จะมาอัตโนมัติ)
                    df = pd.DataFrame(rows_data)
                    
                    # 1. สร้าง Header
                    for col_num, col_name in enumerate(df.columns, 1):
                        cell = ws.cell(row=row_idx, column=col_num, value=col_name)
                        cell.fill = column_fill
                        cell.font = Font(bold=True)
                        cell.border = border
                        cell.alignment = Alignment(horizontal="center")
                    
                    # 2. ใส่ Data ทีละแถว
                    for r in df.itertuples(index=False):
                        row_idx += 1
                        for col_num, val in enumerate(r, 1):
                            cell = ws.cell(row=row_idx, column=col_num, value=val)
                            cell.border = border
                    row_idx += 3
                else:
                    ws.cell(row=row_idx, column=1, value="ไม่พบข้อมูลใน Table นี้")
                    row_idx += 3

        # ปรับความกว้าง Column อัตโนมัติ
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output