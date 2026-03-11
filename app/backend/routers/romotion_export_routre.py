from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import select
from typing import List, Optional,Dict,Any
from datetime import datetime
from fastapi.responses import StreamingResponse, JSONResponse
import openpyxl
from openpyxl.utils import get_column_letter
import io
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
import pandas as pd
from fastapi.responses import StreamingResponse
from app.backend.database import get_db
from app.backend.auth import get_current_user
from app.backend.models.postgres._base_on import Minformationimport, MPromotionHeader, MPromotionBucketEntity,MFileMaster,MVersionControl
from app.backend.schemas.all_schemas import fiter_export,EntityStatusUpdate, AssignUserRequest
from urllib.parse import quote
router = APIRouter(prefix="/export", tags=["export"])

class PromotionDetailRepository:
    
    @staticmethod
    def get_promotion_detail_from_db(req: fiter_export, db: Session) -> List[Dict[str, Any]]:
        """
        Query ดึงข้อมูลจำลอง View VW_PROMOTION_DETAIL JOIN VW_VERSION_FILE
        """
        query = (
            select(
                MPromotionHeader.pro_code.label("Promotion Code"),
                MPromotionHeader.pro_name.label("Promotion Name"),
                MPromotionHeader.start_date.label("Active From"),
                MPromotionHeader.end_date.label("Active To"),
                MPromotionHeader.limit_tran.label("LIMIT"),
                MPromotionBucketEntity.bucket.label("BUCKET"),
                MPromotionBucketEntity.trigger_value.label("TRIGGER"),
                MPromotionBucketEntity.mode.label("AttachmentMode"),
                MPromotionBucketEntity.entity_code.label("Entity Code"),
                MPromotionBucketEntity.entity_name.label("Entity Name"),
                MPromotionBucketEntity.barcode.label("BARCODE"),
                MPromotionBucketEntity.barcode.label("BARCODE CODE39"),
                MPromotionHeader.reward_type.label("Reward Type"),
                MPromotionHeader.reward_value.label("Reward Value"),
                MPromotionHeader.notes.label("NOTES"),
                MPromotionBucketEntity.coupon.label("Coupon ID"),
                MPromotionBucketEntity.coupon.label("COUPON CODE39"),
                MPromotionHeader.reward_ma.label("Reward MA ID"),
                MPromotionHeader.reward_name.label("Reward MA Name"),
                MFileMaster.sheet.label("SHEET"),
                MFileMaster.file_name.label("WORKSHEET"),
                MPromotionHeader.update_date.label("OPTIMAL_DATE"),
                MVersionControl.sr_no.label("VERSION_CONT"),
                Minformationimport.description.label("SYSTEM") 
            )
            .select_from(MPromotionHeader)
            .join(MPromotionBucketEntity, MPromotionHeader.id == MPromotionBucketEntity.pro_id)
            .join(MFileMaster, MPromotionHeader.file_id == MFileMaster.id)
            .join(Minformationimport, MFileMaster.v_id == Minformationimport.id)
            .join(MVersionControl, Minformationimport.v_id == MVersionControl.id)
        )
        if 0 not in req.version_id : 
            query = query.where(MFileMaster.v_id.in_(req.version_id))
            
        elif 0 not in req.file_id :
            query = query.where(MFileMaster.id.in_(req.file_id))
            
        else:
            return []

        result = db.execute(query).mappings().all()
        return [dict(row) for row in result]

TABEL_LINE_LEFT = "thin"
TABEL_LINE_RIGHT = "thin"
TABEL_LINE_TOP = "thin"
TABEL_LINE_BOTTOM = "thin"
FONT_BARCODE = "Bar-Code 39"          # ชื่อฟอนต์ Barcode (สมมติ)
FONT_NOMAL = "Anuphan"
BACKGROUND_FONT_NOMAL = "Anuphan"
HEARDER_FONT_SIZE = 30
BACKGROUND_FONT_SIZE = 12
BACKGROUND_COLOR = "000000"      # สีตัวอักษร Header
HEARDER_COLOR = "DDDDDD"         # สีพื้นหลัง Header
SOLID = "solid"
BACKGROUND_WRAP_TEXT = True
BACKGROUND_HORIZONTAL = "center"
BACKGROUND_VERTICAL = "center"
COLUMN_BARCODE_NAMES = ['BARCODE CODE39', 'COUPON CODE39'] 
PRINT_SIZE_LEFT = 0.25     #1.9
PRINT_SIZE_RIGHT = 0.25     #1.9
PRINT_SIZE_TOP = 0.75     #2.5
PRINT_SIZE_BOTTOM = 0.75     #2.5
PRINT_SIZE_HEADER = 0.30     #1.3
PRINT_SIZE_FOOTER = 0.30     #1.3
PRINT_WIDTH = 1
PRINT_HEIGT = 0
PRINT_SCALE = 100
WIDTH_COLUMN=[10,10,6.3,6.3,7.5,5.3,5.3,5.3,11,11,16.5,53,7.5,10,50,15.3,53,8.6,15,15,15,15]
PRINT_TEXT_HEARDER_RIGHT = "&[Page] of &[Pages]"
PRINT_TEXT_HEARDER_CENTER = "&[Tab]"
# --- ฟังก์ชันทำความสะอาดข้อมูล ---
def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    # เติมค่าว่างด้วย string เปล่า และลบช่องว่างส่วนเกิน
    df = df.fillna("")
    df = df.map(lambda x: str(x).strip() if isinstance(x, str) else x)
    return df

@router.post("/promotion-detail", tags=["Export"])
def export_promotion_details(req:fiter_export,db: Session = Depends(get_db)):
    try:
        # 1. ดึงข้อมูล Data จาก DB (ตามรูปแบบ SELECT ที่กำหนด)
        data = PromotionDetailRepository.get_promotion_detail_from_db(req,db)
        print(len(data))
        if not data:
            return JSONResponse(status_code=403, content={"success": False, "message": "No data found."})

        # โค้ดจัด Format Excel ของคุณ
        df = pd.DataFrame(data, dtype=str)
        df = clean_data(df)
        
        # [FIX] เรียงข้อมูลก่อน เพื่อให้การจัดกลุ่มเส้นขอบทำงานสวยงาม
        sort_cols = [c for c in ["OPTIMAL_DATE", "WORKSHEET"] if c in df.columns]
        if sort_cols:
            df = df.sort_values(by=sort_cols)

        # 2. เตรียม Excel
        output = io.BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Export_Data" # type: ignore

        # 3. Define Styles
        thin_border = Border(
            left=Side(style=TABEL_LINE_LEFT),
            right=Side(style=TABEL_LINE_RIGHT),
            top=Side(style=TABEL_LINE_TOP),
            bottom=Side(style=TABEL_LINE_BOTTOM)
        )
        thick_top_border = Border(
            top=Side(style='thick'),
            left=Side(style=TABEL_LINE_LEFT),
            right=Side(style=TABEL_LINE_RIGHT),
            bottom=Side(style=TABEL_LINE_BOTTOM)
        )
        
        barcode_font = Font(name=FONT_BARCODE, size=HEARDER_FONT_SIZE)
        default_font = Font(name=FONT_NOMAL, size=BACKGROUND_FONT_SIZE)
        header_font = Font(
            name=BACKGROUND_FONT_NOMAL, 
            color=BACKGROUND_COLOR, 
            bold=True, 
            size=BACKGROUND_FONT_SIZE
        )
        
        light_grey_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        header_fill = PatternFill(start_color=HEARDER_COLOR, end_color=HEARDER_COLOR, fill_type=SOLID)
        alignment_style = Alignment(wrap_text=BACKGROUND_WRAP_TEXT, horizontal=BACKGROUND_HORIZONTAL, vertical=BACKGROUND_VERTICAL)

        # 4. Write Data (เขียนครั้งเดียวต่อยาวลงมา)
        headers = list(df.columns)
        
        # เขียน Header ที่บรรทัดที่ 1
        worksheet.append(headers) # type: ignore

        # เขียนข้อมูลทั้งหมด (ต่อยาวลงมา)
        for row in df.itertuples(index=False):
            worksheet.append(list(row)) # type: ignore

        # 5. Formatting Logic (ทำทีเดียวทั้ง Sheet)
        col_map = {name: i+1 for i, name in enumerate(headers)}
        att_mode_idx = col_map.get("AttachmentMode")
        barcode_col_indices = [col_map[c] for c in COLUMN_BARCODE_NAMES if c in col_map]
        
        prev_first_col_val = None
        max_col = worksheet.max_column # type: ignore

        # เริ่ม Loop ตั้งแต่แถว 2 จนจบ
        for r_idx, row_cells in enumerate(worksheet.iter_rows(min_row=2, max_col=max_col), start=2): # type: ignore
            curr_first_col_val = row_cells[0].value
            
            # เช็ค Attachment Mode (สีพื้นหลัง)
            is_exclude = False
            if att_mode_idx:
                att_val = row_cells[att_mode_idx - 1].value
                is_exclude = str(att_val).strip().upper() not in ["INCLUDE",""]

            # Logic เส้นขอบ
            current_border = thin_border
            if r_idx == 2 or curr_first_col_val != prev_first_col_val:
                current_border = thick_top_border
            
            prev_first_col_val = curr_first_col_val

            # Apply Style ราย Cell
            for c_idx, cell in enumerate(row_cells):
                cell_col_num = c_idx + 1
                
                cell.font = default_font
                cell.border = current_border
                cell.alignment = alignment_style

                if is_exclude:
                    cell.fill = light_grey_fill

                if cell_col_num in barcode_col_indices:
                    cell.font = barcode_font

        # 6. Header Style & Widths
        for i, col_name in enumerate(headers):
            cell = worksheet.cell(row=1, column=i+1)  # type: ignore
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = alignment_style
            
            if i < len(WIDTH_COLUMN):
                worksheet.column_dimensions[get_column_letter(i+1)].width = WIDTH_COLUMN[i]  # type: ignore
            else:
                worksheet.column_dimensions[get_column_letter(i+1)].width = 15  # type: ignore

        # 7. Page Setup
        worksheet.row_dimensions[1].height = 105  # type: ignore
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3  # type: ignore
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE  # type: ignore
        worksheet.print_title_rows = "1:1"  # type: ignore
        worksheet.oddHeader.center.text = PRINT_TEXT_HEARDER_CENTER  # type: ignore
        worksheet.oddHeader.right.text = PRINT_TEXT_HEARDER_RIGHT  # type: ignore

        # 8. Save & Return
        workbook.save(output)
        output.seek(0)
        filename = f"Promotion_Export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        headers_http = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
        
        return StreamingResponse(
            io.BytesIO(output.getvalue()), 
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
            headers=headers_http
        )

    except Exception as e:
        print(f"Export Error: {e}")
        return JSONResponse(status_code=500, content={"success": False, "message": str(e)})
