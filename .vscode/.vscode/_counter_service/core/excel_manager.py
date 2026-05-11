import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Dict, Any, List

class ExcelReportManager:
    @staticmethod
    def generate_config_excel_bytes(fetch_result: Dict[str, Any]) -> io.BytesIO:
        """สำหรับ Export ข้อมูล Config Database ให้ออกมาเป็น Excel โดยอ่านจากโครงสร้าง JSON ใหม่"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Config" # type: ignore
        
        header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        column_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
        
        row_idx = 1
        
        # ดึงส่วน "data" ออกมา ซึ่งตอนนี้เป็น List แน่นอน 100%
        tables_list = fetch_result.get("data", [])
        
        # ป้องกัน Error หาก data ไม่ใช่ List (Data Validation)
        if not isinstance(tables_list, list):
            raise TypeError("Expected 'data' to be a list, but got different format.")
        
        for table_dict in tables_list:
            for table_name, rows_data in table_dict.items():
                # เขียนชื่อ Table เป็น Header ใหญ่
                ws.cell(row=row_idx, column=1, value=table_name) # type: ignore
                ws.cell(row=row_idx, column=1).fill = header_fill # type: ignore
                ws.cell(row=row_idx, column=1).font = Font(bold=True) # type: ignore
                row_idx += 2
                
                # หากมีข้อมูล (rows_data เป็น List ของ Dict) โยนเข้า DataFrame ได้เลย
                if rows_data and isinstance(rows_data, list):
                    df = pd.DataFrame(rows_data)
                    
                    # สร้าง Header คอลัมน์
                    for col_num, col_name in enumerate(df.columns, 1):
                        cell = ws.cell(row=row_idx, column=col_num, value=col_name) # type: ignore
                        cell.fill = column_fill
                        cell.font = Font(bold=True)
                        cell.border = border
                        cell.alignment = Alignment(horizontal="center")
                    
                    # ใส่ข้อมูลทีละแถว
                    for r in df.itertuples(index=False):
                        row_idx += 1
                        for col_num, val in enumerate(r, 1):
                            cell = ws.cell(row=row_idx, column=col_num, value=str(val)) # type: ignore
                            cell.border = border
                    row_idx += 3
                else:
                    ws.cell(row=row_idx, column=1, value="ไม่พบข้อมูลใน Table นี้") # type: ignore
                    row_idx += 3

        # ปรับความกว้าง Column อัตโนมัติ
        for col in ws.columns: # type: ignore
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2 # type: ignore

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output